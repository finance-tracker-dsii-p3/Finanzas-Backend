import os
from django.conf import settings
from django.utils import timezone
from django.core.files import File
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from users.models import User
from notifications.models import Notification
from .models import ExportJob


class BasicDataExporter:
    """
    Servicio básico para exportar datos - versión simplificada para proyecto financiero
    """
    
    def __init__(self, export_job):
        self.export_job = export_job
        self.user_ids = export_job.user_ids
        self.start_date = export_job.start_date
        self.end_date = export_job.end_date
    
    def export_users_data(self):
        """Exporta datos básicos de usuarios"""
        try:
            # Obtener queryset de usuarios
            queryset = User.objects.all()
            
            if self.user_ids:
                queryset = queryset.filter(id__in=self.user_ids)
            
            # Crear archivo Excel
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "Usuarios"
            
            # Headers
            headers = [
                'ID', 'Usuario', 'Email', 'Nombre', 'Apellido', 
                'Identificación', 'Teléfono', 'Rol', 'Verificado', 
                'Activo', 'Fecha Registro', 'Último Login'
            ]
            
            for col_num, header in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Datos
            for row_num, user in enumerate(queryset.order_by('username'), 2):
                worksheet.cell(row=row_num, column=1, value=user.id)
                worksheet.cell(row=row_num, column=2, value=user.username)
                worksheet.cell(row=row_num, column=3, value=user.email)
                worksheet.cell(row=row_num, column=4, value=user.first_name)
                worksheet.cell(row=row_num, column=5, value=user.last_name)
                worksheet.cell(row=row_num, column=6, value=user.identification)
                worksheet.cell(row=row_num, column=7, value=user.phone)
                worksheet.cell(row=row_num, column=8, value=user.get_role_display())
                worksheet.cell(row=row_num, column=9, value="Sí" if user.is_verified else "No")
                worksheet.cell(row=row_num, column=10, value="Sí" if user.is_active else "No")
                worksheet.cell(row=row_num, column=11, value=user.date_joined.strftime('%Y-%m-%d %H:%M') if user.date_joined else '')
                worksheet.cell(row=row_num, column=12, value=user.last_login.strftime('%Y-%m-%d %H:%M') if user.last_login else '')
            
            # Ajustar ancho de columnas
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Guardar archivo
            filename = f"usuarios_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join(settings.MEDIA_ROOT, 'exports', filename)
            
            # Crear directorio si no existe
            os.makedirs(os.path.dirname(filepath), exist_ok=True)
            
            workbook.save(filepath)
            
            # Actualizar export job
            with open(filepath, 'rb') as f:
                self.export_job.file.save(filename, File(f))
                self.export_job.mark_as_completed(file_size=os.path.getsize(filepath))
            
            return True
            
        except Exception as e:
            self.export_job.mark_as_failed(str(e))
            return False
    
    def export_notifications_data(self):
        """Exporta datos básicos de notificaciones"""
        try:
            # Obtener queryset de notificaciones
            queryset = Notification.objects.all().select_related('user')
            
            # Filtrar por fechas si se proporcionan
            if self.start_date:
                queryset = queryset.filter(created_at__date__gte=self.start_date)
            if self.end_date:
                queryset = queryset.filter(created_at__date__lte=self.end_date)
            
            # Crear archivo Excel
            workbook = openpyxl.Workbook()
            worksheet = workbook.active
            worksheet.title = "Notificaciones"
            
            # Headers
            headers = [
                'ID', 'Usuario', 'Tipo', 'Título', 'Mensaje', 
                'Leída', 'Fecha Creación', 'Fecha Lectura'
            ]
            
            for col_num, header in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # Datos
            for row_num, notification in enumerate(queryset.order_by('-created_at'), 2):
                worksheet.cell(row=row_num, column=1, value=notification.id)
                worksheet.cell(row=row_num, column=2, value=notification.user.get_full_name())
                worksheet.cell(row=row_num, column=3, value=notification.get_notification_type_display())
                worksheet.cell(row=row_num, column=4, value=notification.title)
                worksheet.cell(row=row_num, column=5, value=notification.message)
                worksheet.cell(row=row_num, column=6, value="Sí" if notification.read else "No")
                worksheet.cell(row=row_num, column=7, value=notification.created_at.strftime('%Y-%m-%d %H:%M'))
                worksheet.cell(row=row_num, column=8, value=notification.read_timestamp.strftime('%Y-%m-%d %H:%M') if notification.read_timestamp else '')
            
            # Ajustar ancho de columnas
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(50, max(12, max_length + 2))  # Límite de ancho
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Guardar archivo
            filename = f"notificaciones_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = os.path.join(settings.MEDIA_ROOT, 'exports', filename)
            
            # Crear directorio si no existe
            os.makedirs(os.path.dirname(filepath), exist_ok=True)
            
            workbook.save(filepath)
            
            # Actualizar export job
            with open(filepath, 'rb') as f:
                self.export_job.file.save(filename, File(f))
                self.export_job.mark_as_completed(file_size=os.path.getsize(filepath))
            
            return True
            
        except Exception as e:
            self.export_job.mark_as_failed(str(e))
            return False


class ExportService:
    """
    Servicio principal para gestionar exportaciones
    """
    
    @staticmethod
    def process_export_job(export_job_id):
        """Procesa un trabajo de exportación"""
        try:
            export_job = ExportJob.objects.get(id=export_job_id)
            export_job.mark_as_processing()
            
            exporter = BasicDataExporter(export_job)
            
            success = False
            if export_job.export_type == ExportJob.USERS_DATA:
                success = exporter.export_users_data()
            elif export_job.export_type == ExportJob.NOTIFICATIONS_DATA:
                success = exporter.export_notifications_data()
            else:
                # Para otros tipos, crear un archivo básico
                success = exporter.export_users_data()
            
            return success
            
        except Exception as e:
            try:
                export_job = ExportJob.objects.get(id=export_job_id)
                export_job.mark_as_failed(str(e))
            except:
                pass
            return False